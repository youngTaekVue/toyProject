import pandas as pd
from openpyxl.styles import Alignment

# CSV 데이터 로드
df = pd.read_csv('files/sample1.csv', encoding='cp949')

for i in range(10):
    file_path = f"files/test{i}.xlsx"

    # [핵심] pd.ExcelWriter를 사용하여 파일 열기
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # 데이터는 3행(index 2)부터 시작하도록 저장
        df.to_excel(writer, index=False, startrow=2, sheet_name='Sheet1')

        # writer 객체에서 현재 활성화된 시트를 바로 가져옴
        ws = writer.sheets['Sheet1']

        # '병합1' 행 작업
        ws['A1'] = "병합1"
        ws.merge_cells('A1:N1')

        # '병합2' 행 작업
        ws['A2'] = "병합2"
        ws.merge_cells('A2:N2')

        # 스타일 적용 (가운데 맞춤)
        center_style = Alignment(horizontal='center', vertical='center')
        ws['A1'].alignment = center_style
        ws['A2'].alignment = center_style

    # with 문이 끝나면 별도의 save() 없이도 파일이 안전하게 저장됩니다.
    print(f"{file_path} 변환 완료")