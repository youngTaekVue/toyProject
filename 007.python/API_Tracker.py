import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os
import time
from datetime import datetime

class ExcelAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("API Tracker")

        # 1. 창 크기 및 화면 중앙 배치
        window_width, window_height = 500, 640
        scr_width = self.root.winfo_screenwidth()
        scr_height = self.root.winfo_screenheight()
        center_x = int((scr_width / 2) - (window_width / 2))
        center_y = int((scr_height / 2) - (window_height / 2))
        self.root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
        self.root.resizable(False, False)

        self.main_frame = tk.Frame(self.root, padx=20, pady=20)
        self.main_frame.pack(expand=True, fill="both")

        tk.Label(self.main_frame, text="API Tracker", font=("맑은 고딕", 16, "bold"), fg="#2c3e50").pack(pady=(0, 15))

        # 2. 버튼 영역
        btn_frame = tk.Frame(self.main_frame)
        btn_frame.pack(pady=10)
        self.start_btn = tk.Button(btn_frame, text="파일 선택 및 분석", command=self.start_analysis_thread,
                                   bg="#2ecc71", fg="white", font=("맑은 고딕", 11, "bold"), width=18, height=2, cursor="hand2")
        self.start_btn.pack(side="left", padx=5)
        self.reset_btn = tk.Button(btn_frame, text="초기화", command=self.reset_all,
                                   bg="#e74c3c", fg="white", font=("맑은 고딕", 11, "bold"), width=10, height=2, cursor="hand2")
        self.reset_btn.pack(side="left", padx=5)

        # 3. 진행 상황 영역
        self.progress_frame = tk.Frame(self.main_frame)
        self.timer_label = tk.Label(self.progress_frame, text="진행 시간: 0초", font=("맑은 고딕", 10, "bold"), fg="#e67e22")
        self.timer_label.pack()
        self.status_label = tk.Label(self.progress_frame, text="준비 중...", fg="#555")
        self.status_label.pack()
        self.progress_bar = ttk.Progressbar(self.progress_frame, orient='horizontal', length=400, mode='determinate')
        self.progress_bar.pack(pady=5)
        self.percent_label = tk.Label(self.progress_frame, text="0%", font=("맑은 고딕", 9, "bold"))
        self.percent_label.pack()

        # 4. 결과 표시 영역
        tk.Label(self.main_frame, text="[ 분석 결과 ]", font=("맑은 고딕", 10, "bold")).pack(pady=(20, 5), anchor="w")
        self.result_area = tk.Text(self.main_frame, height=12, font=("나눔고딕코딩", 10), bg="#f8f9fa", padx=10, pady=10)
        self.result_area.pack(fill="both", expand=True)
        self.set_result_text("파일을 선택하여 분석을 진행해주세요.")

        self.analysis_running = False

    def set_result_text(self, text):
        self.result_area.config(state="normal")
        self.result_area.delete("1.0", tk.END)

        # 기본 스타일 설정
        self.result_area.tag_configure("header", foreground="#2c3e50", font=("맑은 고딕", 11, "bold"))
        self.result_area.tag_configure("success", foreground="#27ae60", font=("맑은 고딕", 12, "bold"))
        self.result_area.tag_configure("fail", foreground="#e74c3c", font=("맑은 고딕", 12, "bold"))
        self.result_area.tag_configure("line", foreground="#bdc3c7")

        self.result_area.insert(tk.END, text)

        # 특정 문구에 색상 입히기 (자동화)
        # 1. 제목 색상
        self.result_area.tag_add("header", "1.0", "1.end")

        # 2. 성공률/실패율 색상 강조
        content = self.result_area.get("1.0", tk.END)
        if "성공률" in content:
            start_idx = content.find("성공률")
            line_num = content.count('\n', 0, start_idx) + 1
            self.result_area.tag_add("success", f"{line_num}.0", f"{line_num}.end")

        if "실패율" in content:
            start_idx = content.find("실패율")
            line_num = content.count('\n', 0, start_idx) + 1
            self.result_area.tag_add("fail", f"{line_num}.0", f"{line_num}.end")

        self.result_area.config(state="disabled")

    def update_timer(self, start_time):
        if self.analysis_running:
            elapsed = int(time.time() - start_time)
            m, s = divmod(elapsed, 60)
            time_str = f"진행 시간: {m}분 {s}초" if m > 0 else f"진행 시간: {s}초"
            self.timer_label.config(text=time_str)
            self.root.after(1000, lambda: self.update_timer(start_time))

    def update_ui_state(self, is_running):
        self.analysis_running = is_running
        if is_running:
            # 분석 중일 때: 모든 버튼 비활성화
            self.start_btn.config(state="disabled", bg="#bdc3c7")
            self.reset_btn.config(state="disabled", bg="#bdc3c7")
            self.progress_frame.pack(pady=10)
        else:
            # 분석이 끝났을 때:
            # '파일 선택' 버튼은 여전히 비활성화 유지, '초기화' 버튼만 활성화
            self.start_btn.config(state="disabled", bg="#bdc3c7")
            self.reset_btn.config(state="normal", bg="#e74c3c")
            # 진행 상태 프레임은 결과 확인을 위해 유지하거나 pack_forget 처리
            # self.progress_frame.pack_forget()

    def reset_all(self):
        if messagebox.askyesno("초기화", "모든 분석 결과와 진행 상황을 초기화하시겠습니까?"):
            self.start_btn.config(state="normal", bg="#2ecc71")
            self.progress_bar['value'] = 0
            self.percent_label.config(text="0%")
            self.status_label.config(text="준비 중...")
            self.timer_label.config(text="진행 시간: 0초")
            self.progress_frame.pack_forget()
            self.set_result_text("파일을 선택하여 분석을 시작해주세요.")


    def perform_analysis(self, files):
        start_time = time.time()
        self.root.after(0, lambda: self.update_timer(start_time))

        total_rows = 0
        filtered_count = 0
        y_count = 0
        n_count = 0

        try:
            total_files = len(files)
            for i, file_path in enumerate(files):
                file_name = os.path.basename(file_path)
                # [추가] 파일 이름이 길 경우 줄이기 (예: 30자 기준)
                display_name = file_name
                if len(display_name) > 30:
                    # 앞 15자 + ... + 뒤 10자 형태로 조합
                    display_name = display_name[:15] + "..." + display_name[-10:]
                    
                self.status_label.config(text=f"분석 중 ({i+1}/{total_files}): {display_name}")

                # openpyxl로 엑셀 읽기 (read_only 모드로 속도 최적화)
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active # 첫 번째 시트 선택

                # 헤더 위치 파악
                headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=3, max_row=3))]
                try:
                    server_idx = headers.index('서버구분')
                    api_idx = headers.index('API통신성공여부')
                except ValueError:
                    continue # 컬럼이 없는 파일은 건너뜀

                # 데이터 순회 (2행부터)
                for row in ws.iter_rows(min_row=4, values_only=True):
                    total_rows += 1

                    # 서버구분이 '1'인 데이터 필터링
                    if str(row[server_idx]) == '1':
                        filtered_count += 1
                        val = str(row[api_idx]) if row[api_idx] else ""
                        if 'Y' in val:
                            y_count += 1
                        elif 'N' in val:
                            n_count += 1

                # 진행률 업데이트
                current_percent = int(((i + 1) / total_files) * 100)
                self.progress_bar['value'] = current_percent
                self.percent_label.config(text=f"{current_percent}%")
                self.root.update_idletasks()
                wb.close()

            # 결과 계산
            y_percent = (y_count / filtered_count * 100) if filtered_count > 0 else 0
            error_percent = (n_count / filtered_count * 100) if filtered_count > 0 else 0

            duration = time.time() - start_time
            m, s = divmod(int(duration), 60)
            time_str = f"{m}분 {s}초" if m > 0 else f"{s}초"

            # perform_analysis 함수 내 result_text 부분 수정
            result_text = (
                f" [ 분석 요약 보고서 ] \n"
                f" 완료 시각 : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f" 실행 시간 : {time_str} ({total_files}개 파일)\n"
                f"{'='*45}\n\n"
                f" 1. 데이터 통계\n"
                f"    • 전체 로드 행     : {total_rows:>10,} 개\n"
                f"    • 필터링 행(구분1) : {filtered_count:>10,} 개\n\n"
                f" 2. API 통신 상세 결과\n"
                f"    • 성공(Y)          : {y_count:>10,} 개\n"
                f"    • 실패(N)          : {n_count:>10,} 개\n\n"
                f" [ 핵심 지표 ]\n"
                f" ──────────────────────────────\n"
                f"   성공률 : {y_percent:>8.2f} %\n"
                f"   실패율 : {error_percent:>8.2f} %\n"
                f" ──────────────────────────────\n"
            )
            self.set_result_text(result_text)

        except Exception as e:
            messagebox.showerror("오류", f"분석 중 오류 발생:\n{e}")
        finally:
            self.update_ui_state(False)

    def start_analysis_thread(self):
        files = filedialog.askopenfilenames(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx")])
        if not files: return
        self.update_ui_state(True)
        threading.Thread(target=self.perform_analysis, args=(files,), daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelAnalyzer(root)
    root.mainloop()